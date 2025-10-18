from __future__ import annotations

import re
import tempfile
import uuid
from html import unescape
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence


def _require_openpyxl() -> tuple[Any, Any]:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Font  # type: ignore
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError("Пакет openpyxl не установлен, Excel-экспорт недоступен") from exc
    return Workbook, {"Alignment": Alignment, "Font": Font}


def _html_to_plain(text: str) -> str:
    if not text:
        return ""
    clean = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    clean = re.sub(r"<[^>]+>", "", clean)
    return unescape(clean).strip()


def _create_temp_path(stem: str) -> Path:
    safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", stem).strip("_") or "export"
    filename = f"{safe_stem}_{uuid.uuid4().hex}.xlsx"
    return Path(tempfile.gettempdir()) / filename


def build_practice_excel(
    *,
    summary_html: str,
    fragments: Sequence[Any],
    structured: Mapping[str, Any] | None = None,
    file_stub: str | None = None,
) -> Path:
    Workbook, styles = _require_openpyxl()
    Alignment = styles["Alignment"]
    Font = styles["Font"]

    wb = Workbook()
    ws_overview = wb.active
    ws_overview.title = "Обзор"

    ws_overview.append(["Раздел", "Содержание"])
    header_font = Font(bold=True)
    for cell in ws_overview[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    summary_text = _html_to_plain(summary_html)
    if structured:
        summary = structured.get("summary") or summary_text
        analysis = structured.get("analysis")
        legal_basis = structured.get("legal_basis") or []
        risks = structured.get("risks") or []
        disclaimer = structured.get("disclaimer")

        ws_overview.append(["Краткий вывод", summary or summary_text or ""])
        if analysis:
            ws_overview.append(["Анализ", analysis])
        if legal_basis:
            lines = []
            for item in legal_basis:
                if isinstance(item, Mapping):
                    reference = str(item.get("reference") or "").strip()
                    explanation = str(item.get("explanation") or "").strip()
                    if reference and explanation:
                        lines.append(f"{reference} — {explanation}")
                    elif reference:
                        lines.append(reference)
                elif item:
                    lines.append(str(item))
            if lines:
                ws_overview.append(["Правовые основания", "\n".join(lines)])
        if risks:
            risk_lines = [str(r).strip() for r in risks if str(r).strip()]
            if risk_lines:
                ws_overview.append(["Риски", "\n".join(risk_lines)])
        if disclaimer:
            ws_overview.append(["Disclaimer", disclaimer])
    else:
        ws_overview.append(["Краткий вывод", summary_text or ""])

    for row in ws_overview.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_overview.column_dimensions["A"].width = 28
    ws_overview.column_dimensions["B"].width = 120

    ws_cases = wb.create_sheet("Практика")
    case_headers = [
        "№",
        "Название / номер / ссылка",
        "Суть дела / обстоятельства",
        "Решение / выводы суда",
        "Нормы, на которые ссылались",
        "Применимость для текущего запроса",
        "Суд",
        "Дата",
        "Регион",
        "Оценка релевантности",
    ]
    ws_cases.append(case_headers)
    for cell in ws_cases[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    if fragments:
        for idx, fragment in enumerate(fragments, start=1):
            match = getattr(fragment, "match", None)
            metadata = getattr(match, "metadata", {}) if match else {}
            if not isinstance(metadata, Mapping):
                metadata = {}
            title = metadata.get("title") or metadata.get("name") or getattr(fragment, "header", "")
            case_number = metadata.get("case_number") or metadata.get("case")
            url = metadata.get("url") or metadata.get("link")
            summary = metadata.get("summary") or getattr(fragment, "excerpt", "")
            decision = metadata.get("decision_summary") or ""
            norms = metadata.get("norms_summary") or ""
            if not norms:
                norm_names = metadata.get("norm_names")
                if isinstance(norm_names, Sequence):
                    norms = "\n".join(str(item).strip() for item in norm_names if str(item).strip())
            applicability = metadata.get("applicability") or ""
            court = metadata.get("court") or ""
            date = metadata.get("date") or metadata.get("decision_date") or ""
            region = metadata.get("region") or ""
            relevance_raw = metadata.get("score")
            if relevance_raw in (None, ""):
                relevance_raw = getattr(match, "score", "")
            if isinstance(relevance_raw, (int, float)):
                relevance_value = f"{float(relevance_raw):.2f}"
            else:
                relevance_value = str(relevance_raw or "").strip()

            name_parts = [str(title or "").strip()]
            if case_number:
                name_parts.append(str(case_number).strip())
            if url:
                name_parts.append(str(url).strip())
            name_cell_value = "\n".join(part for part in name_parts if part)

            ws_cases.append(
                [
                    idx,
                    name_cell_value,
                    str(summary or "").strip(),
                    str(decision or "").strip(),
                    str(norms or "").strip(),
                    str(applicability or "").strip(),
                    str(court or "").strip(),
                    str(date or "").strip(),
                    str(region or "").strip(),
                    relevance_value,
                ]
            )
            if url:
                row_idx = ws_cases.max_row
                link_cell = ws_cases.cell(row=row_idx, column=2)
                link_cell.hyperlink = url
                link_cell.style = "Hyperlink"
    else:
        ws_cases.append([None, "Нет данных по судебной практике", "", "", "", "", "", "", "", ""])

    column_widths = {
        "A": 6,
        "B": 48,
        "C": 60,
        "D": 55,
        "E": 40,
        "F": 55,
        "G": 28,
        "H": 16,
        "I": 28,
        "J": 18,
    }
    for letter, width in column_widths.items():
        ws_cases.column_dimensions[letter].width = width
    for row in ws_cases.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output_path = _create_temp_path(file_stub or "practice")
    wb.save(output_path)
    return output_path


def _format_law_refs(refs: Iterable[Any]) -> str:
    return "\n".join(str(ref).strip() for ref in refs if str(ref).strip())


def build_risk_excel(
    report: Mapping[str, Any],
    *,
    file_stub: str | None = None,
) -> Path:
    Workbook, styles = _require_openpyxl()
    Alignment = styles["Alignment"]
    Font = styles["Font"]

    wb = Workbook()
    ws_overview = wb.active
    ws_overview.title = "Обзор"
    headers = ["Показатель", "Значение"]
    ws_overview.append(headers)
    for cell in ws_overview[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_overview.append(["Общий уровень риска", str(report.get("overall_risk_level") or "")])
    ai_analysis = report.get("ai_analysis") or {}
    ws_overview.append(["ИИ-резюме", str(ai_analysis.get("summary") or "")])
    ws_overview.append(["ИИ-оценка уровня", str(ai_analysis.get("overall_level") or "")])
    ws_overview.append(["Метод анализа", str(ai_analysis.get("method") or "")])
    ws_overview.append(["Количество блоков", str(ai_analysis.get("chunks_analyzed") or "")])

    recommendations = report.get("recommendations") or []
    if recommendations:
        ws_overview.append(["Рекомендации", "\n".join(str(r).strip() for r in recommendations if str(r).strip())])

    ws_overview.column_dimensions["A"].width = 35
    ws_overview.column_dimensions["B"].width = 120
    for row in ws_overview.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _build_risk_sheet(title: str, risks: Iterable[Mapping[str, Any]]) -> None:
        sheet = wb.create_sheet(title)
        sheet.append([
            "ID",
            "Уровень",
            "Описание",
            "Фрагмент договора",
            "Начало",
            "Окончание",
            "Источники права",
            "Источник",
        ])
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        has_rows = False
        for risk in risks or []:
            if not isinstance(risk, Mapping):
                continue
            has_rows = True
            sheet.append([
                str(risk.get("id") or ""),
                str(risk.get("risk_level") or risk.get("level") or ""),
                str(risk.get("description") or ""),
                str(risk.get("clause_text") or ""),
                risk.get("start"),
                risk.get("end"),
                _format_law_refs(risk.get("law_refs") or []),
                str(risk.get("source") or ""),
            ])
        if not has_rows:
            sheet.append(["", "", "Данные отсутствуют", "", "", "", "", ""])
        for letter in ["B", "C", "D", "G", "H"]:
            sheet.column_dimensions[letter].width = 40
        sheet.column_dimensions["E"].width = 12
        sheet.column_dimensions["F"].width = 12
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    _build_risk_sheet("ИИ-риски", ai_analysis.get("risks") or [])
    _build_risk_sheet("Шаблоны", report.get("pattern_risks") or [])

    compliance = report.get("legal_compliance") or {}
    violations = compliance.get("violations") or []
    sheet = wb.create_sheet("Нарушения")
    sheet.append(["ID", "Текст", "Начало", "Окончание", "Источники права", "Комментарий"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    if violations:
        for violation in violations:
            if not isinstance(violation, Mapping):
                continue
            span = violation.get("span") or {}
            sheet.append([
                str(violation.get("id") or ""),
                str(violation.get("text") or ""),
                span.get("start"),
                span.get("end"),
                _format_law_refs(violation.get("law_refs") or []),
                str(violation.get("note") or ""),
            ])
    else:
        sheet.append(["", "Нарушения не обнаружены", "", "", "", ""])
    for letter in ["B", "E", "F"]:
        sheet.column_dimensions[letter].width = 45
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 12
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output_path = _create_temp_path(file_stub or "contract_risks")
    wb.save(output_path)
    return output_path
